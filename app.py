from flask import Flask, request, jsonify
from openpyxl import load_workbook
from werkzeug.utils import secure_filename
import os
import requests

app = Flask(__name__)

UPLOAD_FOLDER = "uploads"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
ALLOWED_EXTENSIONS = {"xlsx", "xls"}

def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route("/upload", methods=["POST"])
def upload_file():
    try:
        # Get dealId from URL
        deal_id = request.args.get("dealId")
        if not deal_id:
            return jsonify({"error": "Missing dealId", "success": False}), 400

        # Collect all uploaded files regardless of key
        files = list(request.files.values())
        if not files:
            return jsonify({"error": "No file received", "success": False}), 400

        result_data = {}

        for file in files:
            if file and allowed_file(file.filename):
                # Save the file locally
                filename = secure_filename(file.filename)
                save_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                file.save(save_path)

                # Open workbook
                workbook = load_workbook(save_path, data_only=True, read_only=True)

                # --- Process Label sheet only ---
                label_sheet = workbook["Label"] if "Label" in workbook.sheetnames else None
                label_data = {
                    "Label_T": {},
                    "Label_TP": {},
                    "Label_C": {},
                    "Label_L": {},
                    "Label_CP": {},
                    "Label_RT": {},
                    "Label_RB": {},
                    "Label_RP": {}
                }
                if label_sheet:
                    row_index = {
                        "T": 1, "TP": 1, "C": 1, "L": 1,
                        "CP": 1, "RT": 1, "RB": 1, "RP": 1
                    }
                    for row in label_sheet.iter_rows(min_row=2):
                        # Column A & B -> Label_T
                        label_data["Label_T"][str(row_index["T"])] = {
                            "T1": row[0].value if len(row) > 0 else None,
                            "T2": row[1].value if len(row) > 1 else None
                        }
                        row_index["T"] += 1

                        # Column C -> Label_TP
                        label_data["Label_TP"][str(row_index["TP"])] = {
                            "TP": row[2].value if len(row) > 2 else None
                        }
                        row_index["TP"] += 1

                        # Column D & E -> Label_C
                        label_data["Label_C"][str(row_index["C"])] = {
                            "C1": row[3].value if len(row) > 3 else None,
                            "C2": row[4].value if len(row) > 4 else None
                        }
                        row_index["C"] += 1

                        # Column F & G -> Label_L
                        label_data["Label_L"][str(row_index["L"])] = {
                            "L1": row[5].value if len(row) > 5 else None,
                            "L2": row[6].value if len(row) > 6 else None
                        }
                        row_index["L"] += 1

                        # Column H -> Label_CP
                        label_data["Label_CP"][str(row_index["CP"])] = {
                            "CP": row[7].value if len(row) > 7 else None
                        }
                        row_index["CP"] += 1

                        # Column I & J -> Label_RT
                        label_data["Label_RT"][str(row_index["RT"])] = {
                            "RT1": row[8].value if len(row) > 8 else None,
                            "RT2": row[9].value if len(row) > 9 else None
                        }
                        row_index["RT"] += 1

                        # Column K & L -> Label_RB
                        label_data["Label_RB"][str(row_index["RB"])] = {
                            "RB1": row[10].value if len(row) > 10 else None,
                            "RB2": row[11].value if len(row) > 11 else None
                        }
                        row_index["RB"] += 1

                        # Column M -> Label_RP
                        label_data["Label_RP"][str(row_index["RP"])] = {
                            "RP": row[12].value if len(row) > 12 else None
                        }
                        row_index["RP"] += 1

                result_data["Label"] = label_data

                # Send data to Zoho CRM
                crm_endpoint = "https://www.zohoapis.com.au/crm/v7/functions/parsedexceldata/actions/execute?auth_type=apikey&zapikey=1003.7aca215a9c900fecfbe589e436532a6a.560d726649a7d347aa1025a35d19c914"
                payload = {
                    "dealId": deal_id,
                    "data": result_data
                }
                headers = {"Content-Type": "application/json"}
                crm_response = requests.post(crm_endpoint, json=payload, headers=headers)

                # Delete the uploaded file
                os.remove(save_path)

            else:
                return jsonify({"error": f"Invalid file type for {file.filename}", "success": False}), 400

        return jsonify({
            "success": True,
            "dealId": deal_id,
            "message": "Files uploaded, processed, sent to CRM, and removed successfully",
            "crm_response": crm_response.json()
        })

    except Exception as e:
        return jsonify({"error": f"Failed to process Excel file: {str(e)}", "success": False}), 500

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
